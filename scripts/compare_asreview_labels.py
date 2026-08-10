#!/usr/bin/env python3
"""
Compare two ASReview label exports and prepare conflicts for adjudication.

Supports:
  - CSV exports from ASReview LAB v3 (asreview_label column)
  - Legacy label_included columns
  - Tag comparison (if tags are present as columns or in notes)
  - .asreview project files (via SQLite extraction)

Usage:
    python compare_asreview_labels.py reviewer_a.csv reviewer_b.csv [output.xlsx]
    python compare_asreview_labels.py reviewer_a.asreview reviewer_b.asreview [output.xlsx]

Output: Excel file with sheets:
  - Summary (statistics: agreement, kappa, tag breakdown)
  - Conflicts (reviewers disagree — for adjudicator)
  - Agreements (both reviewers agree)
  - Unclear (records tagged unclear by either reviewer)
  - No Abstract (records with empty abstracts)
  - Tag Comparison (exclusion reason agreement)
"""

import sys
import csv
import os
import json
import zipfile
import sqlite3
import tempfile
from collections import Counter, defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


# ── Tag definitions (must match asreview-tag-scheme.md) ──────────────

EXCLUSION_TAGS = [
    "not_fnd", "wrong_modality", "not_primary_research", "pediatric_only",
    "case_report", "no_coordinates", "no_human_data", "comorbidity",
    "self_report_dx", "other",
]

WORKFLOW_TAGS = ["unclear", "no_abstract", "needs_fulltext", "prior_knowledge"]

ALL_TAGS = EXCLUSION_TAGS + WORKFLOW_TAGS


# ── Loading ───────────────────────────────────────────────────────────

def load_from_csv(filepath):
    """Load ASReview CSV export. Returns dict keyed by record_id or title."""
    records = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        # Detect delimiter
        first_line = f.readline()
        f.seek(0)
        delim = ";" if first_line.count(";") > first_line.count(",") else ","
        reader = csv.DictReader(f, delimiter=delim)
        for row in reader:
            records.append(row)

    if not records:
        raise ValueError(f"No records found in {filepath}")

    # Determine key field
    key_field = None
    for candidate in ["record_id", "id", "title"]:
        if candidate in records[0] and records[0][candidate].strip():
            key_field = candidate
            break
    if key_field is None:
        raise ValueError(f"Could not find record_id, id, or title column in {filepath}")

    # Determine label field (ASReview v3 uses asreview_label)
    label_field = None
    for candidate in ["asreview_label", "label_included", "included", "label", "relevant"]:
        if candidate in records[0]:
            label_field = candidate
            break
    if label_field is None:
        raise ValueError(f"Could not find label column in {filepath}")

    # Determine notes field
    note_field = None
    for candidate in ["asreview_note", "notes", "note"]:
        if candidate in records[0]:
            note_field = candidate
            break

    result = {}
    for row in records:
        key = row[key_field].strip()
        if not key:
            continue

        # Parse label
        raw_label = row.get(label_field, "").strip()
        try:
            label_int = int(float(raw_label))
        except (ValueError, TypeError):
            label_int = 1 if str(raw_label).lower() in ("true", "yes", "include", "included") else 0

        # Parse tags — look for tag columns or parse from notes
        tags = set()

        # Check for explicit tag columns (ASReview v3 may export tags as columns)
        for col in row:
            col_lower = col.lower()
            if col_lower in ALL_TAGS and row[col].strip():
                tags.add(col_lower)
            elif col_lower == "tags" and row[col].strip():
                # Tags as a single column (comma/semicolon separated)
                for t in row[col].replace(";", ",").split(","):
                    t = t.strip().lower()
                    if t:
                        tags.add(t)

        # Also parse tags from notes (fallback — reviewers may type tags in notes)
        notes = row.get(note_field, "").strip() if note_field else ""
        if notes:
            for tag in ALL_TAGS:
                if tag in notes.lower():
                    tags.add(tag)

        result[key] = {
            "label": label_int,
            "title": row.get("title", "").strip(),
            "abstract": row.get("abstract", "").strip(),
            "notes": notes,
            "tags": tags,
            "raw": row,
        }

    return result, key_field


def load_from_asreview(filepath):
    """Load labels and tags from .asreview project file (SQLite)."""
    # .asreview files are zip archives containing a SQLite database
    with zipfile.ZipFile(filepath, "r") as zf:
        # Find the database file
        db_names = [n for n in zf.namelist() if n.endswith(".sqlite") or "results.db" in n or "project.db" in n]
        if not db_names:
            # Try extracting everything and looking for .db
            with tempfile.TemporaryDirectory() as tmpdir:
                zf.extractall(tmpdir)
                db_files = []
                for root, dirs, files in os.walk(tmpdir):
                    for f in files:
                        if f.endswith(".db") or f.endswith(".sqlite"):
                            db_files.append(os.path.join(root, f))
                if db_files:
                    return _read_sqlite(db_files[0])
            raise ValueError(f"Could not find database in {filepath}")
        # Extract and read
        with tempfile.TemporaryDirectory() as tmpdir:
            zf.extract(db_names[0], tmpdir)
            return _read_sqlite(os.path.join(tmpdir, db_names[0]))


def _read_sqlite(db_path):
    """Read labels and tags from ASReview SQLite database."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # Get all tables
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = [r[0] for r in cursor.fetchall()]

    # Find the record table (usually 'record' or 'records')
    record_table = None
    for candidate in ["record", "records", "article"]:
        if candidate in tables:
            record_table = candidate
            break
    if record_table is None:
        # Try to find a table with 'title' column
        for t in tables:
            try:
                cursor.execute(f"PRAGMA table_info({t})")
                cols = [r[1] for r in cursor.fetchall()]
                if "title" in cols or "abstract" in cols:
                    record_table = t
                    break
            except:
                pass
    if record_table is None:
        raise ValueError(f"Could not find record table in database. Tables: {tables}")

    # Get columns
    cursor.execute(f"PRAGMA table_info({record_table})")
    cols = [r[1] for r in cursor.fetchall()]

    # Find label-related columns
    label_col = None
    for candidate in ["asreview_label", "label_included", "included", "label"]:
        if candidate in cols:
            label_col = candidate
            break

    # Find key column
    key_col = None
    for candidate in ["record_id", "id", "title"]:
        if candidate in cols:
            key_col = candidate
            break

    # Find note column
    note_col = "asreview_note" if "asreview_note" in cols else None

    # Check for tags table (v3)
    tags_table = None
    for candidate in ["tag", "tags", "record_tag"]:
        if candidate in tables:
            tags_table = candidate
            break

    result = {}

    # Load records
    select_cols = [c for c in [key_col, "title", "abstract", label_col, note_col] if c]
    cursor.execute(f"SELECT {', '.join(select_cols)} FROM {record_table}")
    for row in cursor.fetchall():
        key = str(row[key_col]).strip()
        if not key:
            continue
        label_raw = row[label_col] if label_col and label_col in row.keys() else None
        try:
            label = int(label_raw) if label_raw is not None else 0
        except (ValueError, TypeError):
            label = 0

        notes = row[note_col] if note_col and note_col in row.keys() else ""

        result[key] = {
            "label": label,
            "title": row["title"] if "title" in row.keys() else "",
            "abstract": row["abstract"] if "abstract" in row.keys() else "",
            "notes": notes or "",
            "tags": set(),
            "raw": dict(row),
        }

    # Load tags if available
    if tags_table:
        try:
            cursor.execute(f"PRAGMA table_info({tags_table})")
            tag_cols = [r[1] for r in cursor.fetchall()]

            # Find the foreign key to record
            fk_col = None
            for candidate in ["record_id", "id", "article_id"]:
                if candidate in tag_cols:
                    fk_col = candidate
                    break
            tag_name_col = None
            for candidate in ["tag", "name", "tag_name", "value"]:
                if candidate in tag_cols:
                    tag_name_col = candidate
                    break

            if fk_col and tag_name_col:
                cursor.execute(f"SELECT {fk_col}, {tag_name_col} FROM {tags_table}")
                for row in cursor.fetchall():
                    rid = str(row[0]).strip()
                    tag = str(row[1]).strip().lower()
                    if rid in result:
                        result[rid]["tags"].add(tag)
        except Exception:
            pass  # Tags table structure varies, best effort

    conn.close()
    return result, key_col or "record_id"


def load_labels(filepath):
    """Load labels from CSV or .asreview file."""
    if filepath.endswith(".asreview"):
        return load_from_asreview(filepath)
    else:
        return load_from_csv(filepath)


# ── Statistics ────────────────────────────────────────────────────────

def cohen_kappa(a_labels, b_labels):
    """Calculate Cohen's kappa for two lists of binary labels."""
    n = len(a_labels)
    if n == 0:
        return 0.0
    po = sum(1 for a, b in zip(a_labels, b_labels) if a == b) / n
    pa1 = sum(1 for a in a_labels if a == 1) / n
    pb1 = sum(1 for b in b_labels if b == 1) / n
    pe = pa1 * pb1 + (1 - pa1) * (1 - pb1)
    if pe == 1.0:
        return 1.0
    return (po - pe) / (1 - pe)


def kappa_interpretation(kappa):
    if kappa < 0:
        return "Less than chance agreement"
    elif kappa < 0.20:
        return "Slight agreement"
    elif kappa < 0.40:
        return "Fair agreement"
    elif kappa < 0.60:
        return "Moderate agreement"
    elif kappa < 0.80:
        return "Substantial agreement"
    else:
        return "Almost perfect agreement"


# ── Comparison ────────────────────────────────────────────────────────

def compare(file_a, file_b, output_file=None):
    labels_a, key_a = load_labels(file_a)
    labels_b, key_b = load_labels(file_b)

    keys_a = set(labels_a.keys())
    keys_b = set(labels_b.keys())
    common = keys_a & keys_b
    only_a = keys_a - keys_b
    only_b = keys_b - keys_a

    print(f"Reviewer A: {len(keys_a)} records (keyed by '{key_a}')")
    print(f"Reviewer B: {len(keys_b)} records (keyed by '{key_b}')")
    print(f"Common records: {len(common)}")
    if only_a:
        print(f"⚠️  {len(only_a)} records only in A")
    if only_b:
        print(f"⚠️  {len(only_b)} records only in B")
    if not common:
        print("\n❌ No common records. Check that both files use the same record IDs.")
        sys.exit(1)

    agreements = []
    conflicts = []
    unclear_records = []
    no_abstract_records = []
    tag_disagreements = []

    for key in sorted(common):
        a = labels_a[key]
        b = labels_b[key]

        entry = {
            "record_id": key,
            "title": a["title"][:300],
            "abstract": a["abstract"][:500],
            "reviewer_a_label": a["label"],
            "reviewer_b_label": b["label"],
            "reviewer_a_tags": ", ".join(sorted(a["tags"])),
            "reviewer_b_tags": ", ".join(sorted(b["tags"])),
            "reviewer_a_notes": a["notes"][:200],
            "reviewer_b_notes": b["notes"][:200],
            "agree": a["label"] == b["label"],
            "tags_match": a["tags"] == b["tags"],
        }

        # Categorize
        if a["label"] == b["label"]:
            agreements.append(entry)
        else:
            entry["adjudicator_decision"] = ""
            entry["adjudicator_notes"] = ""
            conflicts.append(entry)

        # Check for unclear tag
        if "unclear" in a["tags"] or "unclear" in b["tags"]:
            unclear_records.append(entry)

        # Check for no abstract
        if not a["abstract"].strip():
            no_abstract_records.append(entry)

        # Check for tag disagreement (on exclusion reason specifically)
        if not entry["tags_match"]:
            # Only flag if at least one reviewer had exclusion tags
            a_excl = a["tags"] & set(EXCLUSION_TAGS)
            b_excl = b["tags"] & set(EXCLUSION_TAGS)
            if a_excl or b_excl:
                entry["a_exclusion_tags"] = ", ".join(sorted(a_excl))
                entry["b_exclusion_tags"] = ", ".join(sorted(b_excl))
                tag_disagreements.append(entry)

    # Statistics
    a_labels_list = [labels_a[k]["label"] for k in common]
    b_labels_list = [labels_b[k]["label"] for k in common]
    n = len(common)
    n_agree = len(agreements)
    n_conflict = len(conflicts)
    agreement_rate = n_agree / n if n else 0
    kappa = cohen_kappa(a_labels_list, b_labels_list)
    interp = kappa_interpretation(kappa)

    a_included = sum(a_labels_list)
    b_included = sum(b_labels_list)
    both_included = sum(1 for a, b in zip(a_labels_list, b_labels_list) if a == 1 and b == 1)

    # Tag statistics
    tag_stats = {}
    for tag in ALL_TAGS:
        a_count = sum(1 for k in common if tag in labels_a[k]["tags"])
        b_count = sum(1 for k in common if tag in labels_b[k]["tags"])
        both_count = sum(1 for k in common if tag in labels_a[k]["tags"] and tag in labels_b[k]["tags"])
        tag_stats[tag] = {"a": a_count, "b": b_count, "both": both_count}

    # Print summary
    print(f"\n{'='*60}")
    print(f"SCREENING COMPARISON SUMMARY")
    print(f"{'='*60}")
    print(f"Records compared:        {n}")
    print(f"Agreements:             {n_agree} ({agreement_rate:.1%})")
    print(f"Conflicts:              {n_conflict} ({1-agreement_rate:.1%})")
    print(f"")
    print(f"Reviewer A included:     {a_included}")
    print(f"Reviewer B included:     {b_included}")
    print(f"Both included:           {both_included}")
    print(f"")
    print(f"Cohen's κ:              {kappa:.3f} ({interp})")
    print(f"")
    print(f"Tag statistics:")
    print(f"  {'Tag':<25} {'A':>5} {'B':>5} {'Both':>5}")
    print(f"  {'-'*45}")
    for tag in ALL_TAGS:
        s = tag_stats[tag]
        if s["a"] or s["b"]:
            print(f"  {tag:<25} {s['a']:>5} {s['b']:>5} {s['both']:>5}")
    print(f"")
    print(f"Unclear records:        {len(unclear_records)}")
    print(f"No-abstract records:     {len(no_abstract_records)}")
    print(f"Tag disagreements:      {len(tag_disagreements)}")
    print(f"{'='*60}")

    # Write output
    if output_file is None:
        output_file = os.path.join(
            os.path.dirname(os.path.abspath(file_a)),
            "screening_comparison.xlsx"
        )

    stats = {
        "n_total": n,
        "n_agree": n_agree,
        "n_conflict": n_conflict,
        "agreement_rate": agreement_rate,
        "kappa": kappa,
        "interp": interp,
        "a_included": a_included,
        "b_included": b_included,
        "both_included": both_included,
        "n_unclear": len(unclear_records),
        "n_no_abstract": len(no_abstract_records),
        "n_tag_disagree": len(tag_disagreements),
        "tag_stats": tag_stats,
    }

    if HAS_EXCEL:
        _write_excel(agreements, conflicts, unclear_records,
                     no_abstract_records, tag_disagreements, output_file, stats)
        print(f"\n✅ Excel written: {output_file}")
        print(f"   Sheets: Summary, Conflicts ({n_conflict}), Agreements ({n_agree}),")
        print(f"           Unclear ({len(unclear_records)}), No Abstract ({len(no_abstract_records)}),")
        print(f"           Tag Comparison ({len(tag_disagreements)})")
    else:
        _write_csv_fallback(agreements, conflicts, output_file)
        print(f"\n✅ CSV files written (install openpyxl for Excel output)")

    if n_conflict > 0:
        print(f"\n⚠️  {n_conflict} conflicts need adjudication.")
    if unclear_records:
        print(f"⚠️  {len(unclear_records)} records tagged unclear — review at full text.")
    if no_abstract_records:
        print(f"⚠️  {len(no_abstract_records)} records with no abstract — verify at full text.")

    return stats


# ── Excel output ─────────────────────────────────────────────────────

def _write_excel(agreements, conflicts, unclear, no_abstract,
                 tag_disagree, filepath, stats):
    wb = Workbook()
    bold = Font(bold=True)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws.append(["SCREENING COMPARISON SUMMARY"])
    ws.append([])
    ws.append(["Metric", "Value"])
    for c in ws[3]: c.font = bold

    rows = [
        ("Total records compared", stats["n_total"]),
        ("Agreements", stats["n_agree"]),
        ("Conflicts", stats["n_conflict"]),
        ("Agreement rate", f"{stats['agreement_rate']:.1%}"),
        ("Cohen's κ", f"{stats['kappa']:.3f}"),
        ("Interpretation", stats["interp"]),
        ("", ""),
        ("Reviewer A included", stats["a_included"]),
        ("Reviewer B included", stats["b_included"]),
        ("Both included", stats["both_included"]),
        ("", ""),
        ("Unclear records", stats["n_unclear"]),
        ("No-abstract records", stats["n_no_abstract"]),
        ("Tag disagreements", stats["n_tag_disagree"]),
        ("", ""),
        ("TAG BREAKDOWN", ""),
        ("Tag", "A | B | Both"),
    ]
    for r in rows:
        ws.append(list(r) if isinstance(r, tuple) else [r])

    for tag in ALL_TAGS:
        s = stats["tag_stats"][tag]
        if s["a"] or s["b"]:
            ws.append([tag, f"{s['a']} | {s['b']} | {s['both']}"])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    # ── Conflicts sheet ──
    ws_c = wb.create_sheet("Conflicts")
    headers = [
        "record_id", "title", "abstract",
        "reviewer_a_label", "reviewer_b_label",
        "reviewer_a_tags", "reviewer_b_tags",
        "reviewer_a_notes", "reviewer_b_notes",
        "adjudicator_decision", "adjudicator_notes"
    ]
    ws_c.append(headers)
    for c in ws_c[1]:
        c.font = bold
        c.fill = red_fill
    for e in conflicts:
        ws_c.append([
            e["record_id"], e["title"], e["abstract"][:300],
            e["reviewer_a_label"], e["reviewer_b_label"],
            e["reviewer_a_tags"], e["reviewer_b_tags"],
            e["reviewer_a_notes"], e["reviewer_b_notes"],
            "", ""
        ])
    ws_c.column_dimensions["B"].width = 50
    ws_c.column_dimensions["C"].width = 60
    ws_c.freeze_panes = "A2"

    # ── Agreements sheet ──
    ws_a = wb.create_sheet("Agreements")
    headers = ["record_id", "title", "consensus_label", "reviewer_a_tags", "reviewer_b_tags"]
    ws_a.append(headers)
    for c in ws_a[1]: c.font = bold
    for e in agreements:
        ws_a.append([
            e["record_id"], e["title"][:200],
            e["reviewer_a_label"],
            e["reviewer_a_tags"], e["reviewer_b_tags"],
        ])
    ws_a.column_dimensions["B"].width = 60
    ws_a.freeze_panes = "A2"

    # ── Unclear sheet ──
    ws_u = wb.create_sheet("Unclear")
    headers = [
        "record_id", "title", "abstract",
        "reviewer_a_label", "reviewer_b_label",
        "reviewer_a_tags", "reviewer_b_tags",
        "reviewer_a_notes", "reviewer_b_notes",
        "fulltext_decision", "final_label"
    ]
    ws_u.append(headers)
    for c in ws_u[1]:
        c.font = bold
        c.fill = yellow_fill
    for e in unclear:
        ws_u.append([
            e["record_id"], e["title"], e["abstract"][:300],
            e["reviewer_a_label"], e["reviewer_b_label"],
            e["reviewer_a_tags"], e["reviewer_b_tags"],
            e["reviewer_a_notes"], e["reviewer_b_notes"],
            "", ""
        ])
    ws_u.column_dimensions["B"].width = 50
    ws_u.column_dimensions["C"].width = 60
    ws_u.freeze_panes = "A2"

    # ── No Abstract sheet ──
    ws_na = wb.create_sheet("No Abstract")
    headers = [
        "record_id", "title",
        "reviewer_a_label", "reviewer_b_label",
        "reviewer_a_tags", "reviewer_b_tags",
        "fulltext_decision", "final_label"
    ]
    ws_na.append(headers)
    for c in ws_na[1]:
        c.font = bold
        c.fill = yellow_fill
    for e in no_abstract:
        ws_na.append([
            e["record_id"], e["title"],
            e["reviewer_a_label"], e["reviewer_b_label"],
            e["reviewer_a_tags"], e["reviewer_b_tags"],
            "", ""
        ])
    ws_na.column_dimensions["B"].width = 70
    ws_na.freeze_panes = "A2"

    # ── Tag Comparison sheet ──
    ws_t = wb.create_sheet("Tag Comparison")
    headers = [
        "record_id", "title",
        "agree_on_label?",
        "reviewer_a_tags", "reviewer_b_tags",
        "a_exclusion_tags", "b_exclusion_tags",
        "tag_mismatch_notes"
    ]
    ws_t.append(headers)
    for c in ws_t[1]:
        c.font = bold
        c.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for e in tag_disagree:
        ws_t.append([
            e["record_id"], e["title"][:200],
            "Yes" if e["agree"] else "No",
            e["reviewer_a_tags"], e["reviewer_b_tags"],
            e.get("a_exclusion_tags", ""), e.get("b_exclusion_tags", ""),
            "",
        ])
    ws_t.column_dimensions["B"].width = 50
    ws_t.freeze_panes = "A2"

    wb.save(filepath)


def _write_csv_fallback(agreements, conflicts, base_path):
    base = base_path.replace(".xlsx", "")
    with open(f"{base}_conflicts.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["record_id", "title", "reviewer_a_label", "reviewer_b_label",
                     "reviewer_a_tags", "reviewer_b_tags", "adjudicator_decision"])
        for e in conflicts:
            w.writerow([e["record_id"], e["title"], e["reviewer_a_label"],
                        e["reviewer_b_label"], e["reviewer_a_tags"], e["reviewer_b_tags"], ""])
    with open(f"{base}_agreements.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["record_id", "title", "consensus_label", "tags_a", "tags_b"])
        for e in agreements:
            w.writerow([e["record_id"], e["title"], e["reviewer_a_label"],
                        e["reviewer_a_tags"], e["reviewer_b_tags"]])


# ── Main ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python compare_asreview_labels.py <reviewer_a> <reviewer_b> [output.xlsx]")
        print("")
        print("  <reviewer_a/b> can be .csv (ASReview export) or .asreview (project file)")
        sys.exit(1)

    a_file = sys.argv[1]
    b_file = sys.argv[2]
    out = sys.argv[3] if len(sys.argv) > 3 else None

    compare(a_file, b_file, out)
